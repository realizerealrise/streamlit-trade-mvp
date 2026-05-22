"""투자성과 대시보드 엑셀 보고서 생성

3개 시트 구조:
1. ★ 투자성과 대시보드 — 요약/지표/TOP·BOTTOM/국가별
2. ★ 종목손익현황 — 종목별 손익 요약표 (대시보드의 데이터 소스)
3. 종목별 거래내역 — 종목당 1시트 (대시보드에서 하이퍼링크)

스타일: 사장님 워크북 양식 재현 (Arial, 파란 헤더, 노란 합계)
"""
import io
import re
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────── 스타일 ───────────────
COLOR_HEADER = '4472C4'        # 파란 헤더
COLOR_HEADER_TEXT = 'FFFFFF'
COLOR_TOTAL = 'FFF2CC'         # 노란 합계
COLOR_SUBTOTAL = 'F2F2F2'      # 회색 소계
COLOR_TITLE = '1F3864'         # 진한 파랑 (제목)
COLOR_SECTION_BORDER = '4472C4'
COLOR_LINK = '0563C1'
COLOR_BORDER = 'BFBFBF'

FONT_TITLE = Font(name='Arial', size=14, bold=True, color=COLOR_TITLE)
FONT_SUBTITLE = Font(name='Arial', size=9, color='666666')
FONT_SECTION = Font(name='Arial', size=11, bold=True, color=COLOR_TITLE)
FONT_HEADER = Font(name='Arial', size=10, bold=True, color=COLOR_HEADER_TEXT)
FONT_BODY = Font(name='Arial', size=10)
FONT_BODY_BOLD = Font(name='Arial', size=10, bold=True)
FONT_LINK = Font(name='Arial', size=10, color=COLOR_LINK, underline='single')

FILL_HEADER = PatternFill('solid', fgColor=COLOR_HEADER)
FILL_TOTAL = PatternFill('solid', fgColor=COLOR_TOTAL)
FILL_SUBTOTAL = PatternFill('solid', fgColor=COLOR_SUBTOTAL)

THIN = Side(style='thin', color=COLOR_BORDER)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', indent=1)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)

# 숫자 포맷 (양수, 음수, 0)
FMT_KRW = '#,##0;(#,##0);"-"'
FMT_PCT = '0.0%;(0.0%);"-"'
FMT_INT = '#,##0;(#,##0);"-"'


# ─────────────── 시트명 정리 ───────────────
INVALID_SHEET_CHARS = re.compile(r'[\\/?*\[\]:]')


def _safe_sheet_name(name, used_names):
    """Excel 시트명 제약 처리 (31자, 특수문자 제거, 중복 방지)"""
    s = INVALID_SHEET_CHARS.sub('_', str(name)).strip()
    s = s[:28] if len(s) > 28 else s  # 여유 두기 (중복 시 _N 추가용)
    base = s if s else 'Sheet'
    
    final = base
    i = 1
    while final in used_names or final.lower() in {x.lower() for x in used_names}:
        suffix = f'_{i}'
        final = (base[:28-len(suffix)] + suffix)
        i += 1
    used_names.add(final)
    return final


# ─────────────── 메인 진입점 ───────────────
def build_dashboard_workbook(trades, pnl_df, holdings_df, current_prices=None, fx_rates=None, year=None):
    """
    투자성과 대시보드 워크북 생성
    
    Args:
        trades: list of dict — 통합 거래 내역
        pnl_df: DataFrame — 종목별 손익 (calculate_stock_pnl 결과)
        holdings_df: DataFrame — 보유 종목 (get_current_holdings 결과)
        current_prices: dict — 현재가 (선택)
        fx_rates: dict — 환율 (선택)
        year: int — 대상 연도 (None이면 거래에서 자동 감지)
    
    Returns:
        BytesIO — 엑셀 파일 바이너리
    """
    if year is None:
        years = set()
        for t in trades:
            try:
                years.add(int(t['거래일자'][:4]))
            except (ValueError, KeyError, IndexError):
                pass
        year = max(years) if years else datetime.now().year
    
    wb = Workbook()
    # 기본 시트 제거
    wb.remove(wb.active)
    
    used_names = set()
    
    # 종목명 → 시트명 매핑 (하이퍼링크용)
    stock_sheet_map = {}
    if not pnl_df.empty:
        for _, row in pnl_df.iterrows():
            sheet_name = _safe_sheet_name(row['종목명'] or '기타', used_names)
            stock_sheet_map[row['종목명']] = sheet_name
    
    # 1. 대시보드 (먼저 만들지만 데이터는 종목손익현황 참조라 나중에 채움)
    ws_dash = wb.create_sheet('★투자성과 대시보드')
    used_names.add('★투자성과 대시보드')
    
    # 2. 종목손익현황
    ws_pnl = wb.create_sheet('★종목손익현황')
    used_names.add('★종목손익현황')
    _build_stock_pnl_sheet(ws_pnl, pnl_df, holdings_df, stock_sheet_map, trades=trades)
    
    # 3. 종목별 거래내역 시트
    for stock_name, sheet_name in stock_sheet_map.items():
        ws_stock = wb.create_sheet(sheet_name)
        _build_stock_detail_sheet(ws_stock, stock_name, trades, pnl_df, holdings_df)
    
    # 1. 대시보드 마지막에 채움 (수식 참조)
    _build_dashboard_sheet(ws_dash, pnl_df, holdings_df, year, stock_sheet_map)
    
    # 시트 순서 정리: 대시보드 → 종목손익현황 → 종목 시트들
    desired_order = ['★투자성과 대시보드', '★종목손익현황']
    desired_order.extend([stock_sheet_map[s] for s in stock_sheet_map])
    
    # openpyxl의 _sheets 직접 재정렬
    current = {ws.title: ws for ws in wb.worksheets}
    new_order = [current[name] for name in desired_order if name in current]
    # 누락된 시트가 있으면 끝에 추가
    new_order.extend([ws for ws in wb.worksheets if ws not in new_order])
    wb._sheets = new_order
    
    # BytesIO 출력
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ─────────────── 시트 1: ★투자성과 대시보드 ───────────────
def _build_dashboard_sheet(ws, pnl_df, holdings_df, year, stock_sheet_map):
    """대시보드 시트 빌드 — 사장님 양식 재현"""
    ws.sheet_view.showGridLines = False
    
    # 열 너비
    col_widths = {'A': 1, 'B': 30, 'C': 22, 'D': 18, 'E': 18, 'F': 18, 'G': 14, 'H': 2,
                  'I': 28, 'J': 18, 'K': 18, 'L': 18}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    
    pnl_sheet = "'★종목손익현황'"
    last_row = len(pnl_df) + 3 if not pnl_df.empty else 4
    total_row = last_row + 1
    
    # ─── 제목 ───
    ws['B1'] = f'★ {year}년 투자 성과 대시보드'
    ws['B1'].font = FONT_TITLE
    ws.merge_cells('B1:G1')
    
    ws['B2'] = f'분석 기준: {datetime.now().strftime("%Y-%m-%d")}  ·  토스 / 한투 / 나무 통합'
    ws['B2'].font = FONT_SUBTITLE
    ws.merge_cells('B2:G2')
    
    # ─── 섹션 1: 연간 투자 성과 요약 (B4:C12) ───
    ws['B4'] = '■ 연간 투자 성과 요약'
    ws['B4'].font = FONT_SECTION
    
    summary_data = [
        ('총 매수금액', f"=SUM({pnl_sheet}!E4:E{last_row})"),
        ('총 매도금액', f"=SUM({pnl_sheet}!H4:H{last_row})"),
        ('기말 보유잔고(평가)', f"=SUM({pnl_sheet}!N4:N{last_row})"),
        ('총 투자종목 수', f"=COUNTA({pnl_sheet}!B4:B{last_row})"),
        ('보유 종목 수', f"=COUNTIF({pnl_sheet}!M4:M{last_row},\">0\")"),
        ('매도 완료 종목 수', f"=COUNTIF({pnl_sheet}!I4:I{last_row},\">0\")"),
    ]
    _write_kv_table(ws, 5, 'B', '항목', '금액', summary_data, fmt_value=FMT_KRW)
    
    # ─── 섹션 2: 손익 분석 (E4:F11) ───
    ws['E4'] = '■ 손익 분석'
    ws['E4'].font = FONT_SECTION
    
    pnl_data = [
        ('매도차손익', f"=SUM({pnl_sheet}!J4:J{last_row})"),
        ('처분이익 (+)', f"=SUM({pnl_sheet}!K4:K{last_row})"),
        ('처분손실 (−)', f"=SUM({pnl_sheet}!L4:L{last_row})"),
        ('수수료 합계', f"=SUM({pnl_sheet}!O4:O{last_row})"),
        ('거래세 합계', f"=SUM({pnl_sheet}!P4:P{last_row})"),
        ('투자 순수익 (매도차손익 − 비용)', f"=F6-F9-F10"),
    ]
    _write_kv_table(ws, 5, 'E', '항목', '금액', pnl_data, fmt_value=FMT_KRW, last_row_total=True)
    
    # ─── 섹션 3: 핵심 투자 지표 (I4:J9) ───
    ws['I4'] = '■ 핵심 투자 지표'
    ws['I4'].font = FONT_SECTION
    
    kpi_data = [
        ('매도실현 수익률', f"=IFERROR(SUM({pnl_sheet}!J4:J{last_row})/SUM({pnl_sheet}!H4:H{last_row}),0)", FMT_PCT),
        ('승률 (이익종목 ÷ 매도종목)', f"=IFERROR(COUNTIF({pnl_sheet}!J4:J{last_row},\">0\")/COUNTIF({pnl_sheet}!I4:I{last_row},\">0\"),0)", FMT_PCT),
        ('이익 종목 수', f"=COUNTIF({pnl_sheet}!J4:J{last_row},\">0\")", FMT_INT),
        ('손실 종목 수', f"=COUNTIF({pnl_sheet}!J4:J{last_row},\"<0\")", FMT_INT),
        ('평균 실현이익', f"=IFERROR(SUMIF({pnl_sheet}!J4:J{last_row},\">0\")/COUNTIF({pnl_sheet}!J4:J{last_row},\">0\"),0)", FMT_KRW),
        ('평균 실현손실', f"=IFERROR(SUMIF({pnl_sheet}!J4:J{last_row},\"<0\")/COUNTIF({pnl_sheet}!J4:J{last_row},\"<0\"),0)", FMT_KRW),
    ]
    _write_kpi_table(ws, 5, 'I', '지표', '값', kpi_data)
    
    # ─── 섹션 4: 종목별 실현수익 TOP 10 (B15:G27) ───
    current_row = 15
    ws.cell(row=current_row, column=2, value='■ 종목별 실현수익 TOP 10').font = FONT_SECTION
    current_row += 1
    _write_top_bottom_section(ws, current_row, pnl_df, pnl_sheet, last_row, stock_sheet_map,
                              top=True, by='실현금액')
    current_row += 12
    
    # ─── 섹션 5: 종목별 실현손실 BOTTOM 10 ───
    ws.cell(row=current_row, column=2, value='■ 종목별 실현손실 BOTTOM 10').font = FONT_SECTION
    current_row += 1
    _write_top_bottom_section(ws, current_row, pnl_df, pnl_sheet, last_row, stock_sheet_map,
                              top=False, by='실현금액')
    current_row += 12
    
    # ─── 섹션 6: 수익률 TOP 10 ───
    ws.cell(row=current_row, column=2, value='■ 종목별 수익률 TOP 10').font = FONT_SECTION
    current_row += 1
    _write_top_bottom_section(ws, current_row, pnl_df, pnl_sheet, last_row, stock_sheet_map,
                              top=True, by='수익률')
    current_row += 12
    
    # ─── 섹션 7: 손실률 BOTTOM 10 ───
    ws.cell(row=current_row, column=2, value='■ 종목별 손실률 BOTTOM 10').font = FONT_SECTION
    current_row += 1
    _write_top_bottom_section(ws, current_row, pnl_df, pnl_sheet, last_row, stock_sheet_map,
                              top=False, by='수익률')
    current_row += 12
    
    # ─── 섹션 8: 국가/통화별 손익 (I15~) ───
    ws.cell(row=15, column=9, value='■ 국가/통화별 투자 손익').font = FONT_SECTION
    _write_country_table(ws, 16, 'I', pnl_sheet, last_row)


def _write_kv_table(ws, start_row, start_col, h1, h2, data, fmt_value=FMT_KRW, last_row_total=False):
    """2열 key-value 표 (헤더 + 데이터)"""
    c1 = start_col
    c2 = chr(ord(start_col) + 1)
    
    # 헤더
    ws[f'{c1}{start_row}'] = h1
    ws[f'{c2}{start_row}'] = h2
    for cell_ref in (f'{c1}{start_row}', f'{c2}{start_row}'):
        c = ws[cell_ref]
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    
    # 데이터
    for i, (k, v) in enumerate(data):
        r = start_row + 1 + i
        ws[f'{c1}{r}'] = k
        ws[f'{c2}{r}'] = v
        ws[f'{c1}{r}'].font = FONT_BODY
        ws[f'{c1}{r}'].alignment = ALIGN_LEFT
        ws[f'{c1}{r}'].border = BORDER_ALL
        ws[f'{c2}{r}'].font = FONT_BODY
        ws[f'{c2}{r}'].alignment = ALIGN_RIGHT
        ws[f'{c2}{r}'].number_format = fmt_value
        ws[f'{c2}{r}'].border = BORDER_ALL
        
        # 마지막 행은 합계 강조
        if last_row_total and i == len(data) - 1:
            ws[f'{c1}{r}'].fill = FILL_TOTAL
            ws[f'{c2}{r}'].fill = FILL_TOTAL
            ws[f'{c1}{r}'].font = FONT_BODY_BOLD
            ws[f'{c2}{r}'].font = FONT_BODY_BOLD


def _write_kpi_table(ws, start_row, start_col, h1, h2, data):
    """3열 KPI 표 (헤더 + 데이터 + 포맷 개별)"""
    c1 = start_col
    c2 = chr(ord(start_col) + 1)
    
    ws[f'{c1}{start_row}'] = h1
    ws[f'{c2}{start_row}'] = h2
    for cell_ref in (f'{c1}{start_row}', f'{c2}{start_row}'):
        c = ws[cell_ref]
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    
    for i, (k, v, fmt) in enumerate(data):
        r = start_row + 1 + i
        ws[f'{c1}{r}'] = k
        ws[f'{c2}{r}'] = v
        ws[f'{c1}{r}'].font = FONT_BODY
        ws[f'{c1}{r}'].alignment = ALIGN_LEFT
        ws[f'{c1}{r}'].border = BORDER_ALL
        ws[f'{c2}{r}'].font = FONT_BODY_BOLD
        ws[f'{c2}{r}'].alignment = ALIGN_RIGHT
        ws[f'{c2}{r}'].number_format = fmt
        ws[f'{c2}{r}'].border = BORDER_ALL


def _write_top_bottom_section(ws, start_row, pnl_df, pnl_sheet, last_row, stock_sheet_map,
                               top=True, by='실현금액'):
    """TOP/BOTTOM 10 섹션 - 순위/종목명/매수금액/실현손익/수익률"""
    # 헤더
    headers = ['순위', '종목명', '매수금액', '실현손익', '수익률']
    col_letters = ['B', 'C', 'E', 'F', 'G']
    
    for letter, header in zip(col_letters, headers):
        c = ws[f'{letter}{start_row}']
        c.value = header
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    
    # 데이터 정렬
    if pnl_df.empty:
        return
    
    df = pnl_df.copy()
    # 매도가 있는 종목만 (수익률 의미)
    df = df[df['총매도(원)'] > 0]
    if df.empty:
        return
    
    df['_수익률'] = df['처분손익(원)'] / df['총매수(원)'].replace(0, 1)
    
    if by == '실현금액':
        df = df.sort_values('처분손익(원)', ascending=not top)
    else:  # 수익률
        df = df.sort_values('_수익률', ascending=not top)
    
    top10 = df.head(10)
    
    for i, (_, row) in enumerate(top10.iterrows()):
        r = start_row + 1 + i
        ws[f'B{r}'] = i + 1
        ws[f'B{r}'].font = FONT_BODY
        ws[f'B{r}'].alignment = ALIGN_CENTER
        ws[f'B{r}'].border = BORDER_ALL
        
        # 종목명 + 하이퍼링크
        stock_name = row['종목명']
        sheet_name = stock_sheet_map.get(stock_name, '')
        if sheet_name:
            ws[f'C{r}'] = f"=HYPERLINK(\"#'{sheet_name}'!A1\",\"{stock_name}\")"
            ws[f'C{r}'].font = FONT_LINK
        else:
            ws[f'C{r}'] = stock_name
            ws[f'C{r}'].font = FONT_BODY
        ws[f'C{r}'].alignment = ALIGN_WRAP
        ws[f'C{r}'].border = BORDER_ALL
        ws.merge_cells(f'C{r}:D{r}')
        
        ws[f'E{r}'] = row['총매수(원)']
        ws[f'F{r}'] = row['처분손익(원)']
        ws[f'G{r}'] = f"=IFERROR(F{r}/E{r},0)"
        for letter in ['E', 'F']:
            ws[f'{letter}{r}'].font = FONT_BODY
            ws[f'{letter}{r}'].alignment = ALIGN_RIGHT
            ws[f'{letter}{r}'].number_format = FMT_KRW
            ws[f'{letter}{r}'].border = BORDER_ALL
        ws[f'G{r}'].font = FONT_BODY_BOLD
        ws[f'G{r}'].alignment = ALIGN_RIGHT
        ws[f'G{r}'].number_format = FMT_PCT
        ws[f'G{r}'].border = BORDER_ALL
        
        # 행 높이 (긴 종목명 대응)
        ws.row_dimensions[r].height = 22


def _write_country_table(ws, start_row, start_col, pnl_sheet, last_row):
    """국가/통화별 손익 요약표"""
    c1 = start_col  # I
    c2 = chr(ord(c1) + 1)  # J
    c3 = chr(ord(c1) + 2)  # K
    c4 = chr(ord(c1) + 3)  # L
    
    headers = ['국가(통화)', '종목수', '매수총액', '실현손익']
    cols = [c1, c2, c3, c4]
    
    for letter, h in zip(cols, headers):
        c = ws[f'{letter}{start_row}']
        c.value = h
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    
    countries = [
        ('🇰🇷 한국 (KRW)', 'KRW'),
        ('🇺🇸 미국 (USD)', 'USD'),
        ('🇯🇵 일본 (JPY)', 'JPY'),
        ('🇭🇰 홍콩 (HKD)', 'HKD'),
        ('🇨🇳 중국 (CNY)', 'CNY'),
    ]
    
    for i, (name, currency) in enumerate(countries):
        r = start_row + 1 + i
        ws[f'{c1}{r}'] = name
        ws[f'{c2}{r}'] = f'=COUNTIF({pnl_sheet}!C4:C{last_row},"{currency}")'
        ws[f'{c3}{r}'] = f'=SUMIF({pnl_sheet}!C4:C{last_row},"{currency}",{pnl_sheet}!E4:E{last_row})'
        ws[f'{c4}{r}'] = f'=SUMIF({pnl_sheet}!C4:C{last_row},"{currency}",{pnl_sheet}!J4:J{last_row})'
        
        ws[f'{c1}{r}'].font = FONT_BODY
        ws[f'{c1}{r}'].alignment = ALIGN_LEFT
        ws[f'{c1}{r}'].border = BORDER_ALL
        for letter in [c2, c3, c4]:
            ws[f'{letter}{r}'].font = FONT_BODY
            ws[f'{letter}{r}'].alignment = ALIGN_RIGHT
            ws[f'{letter}{r}'].border = BORDER_ALL
        ws[f'{c2}{r}'].number_format = FMT_INT
        ws[f'{c3}{r}'].number_format = FMT_KRW
        ws[f'{c4}{r}'].number_format = FMT_KRW
    
    # 합계 행
    total_r = start_row + 6
    ws[f'{c1}{total_r}'] = '합계'
    ws[f'{c2}{total_r}'] = f'=SUM({c2}{start_row+1}:{c2}{start_row+5})'
    ws[f'{c3}{total_r}'] = f'=SUM({c3}{start_row+1}:{c3}{start_row+5})'
    ws[f'{c4}{total_r}'] = f'=SUM({c4}{start_row+1}:{c4}{start_row+5})'
    
    for letter in [c1, c2, c3, c4]:
        ws[f'{letter}{total_r}'].fill = FILL_TOTAL
        ws[f'{letter}{total_r}'].font = FONT_BODY_BOLD
        ws[f'{letter}{total_r}'].border = BORDER_ALL
    ws[f'{c1}{total_r}'].alignment = ALIGN_LEFT
    for letter in [c2, c3, c4]:
        ws[f'{letter}{total_r}'].alignment = ALIGN_RIGHT
    ws[f'{c2}{total_r}'].number_format = FMT_INT
    ws[f'{c3}{total_r}'].number_format = FMT_KRW
    ws[f'{c4}{total_r}'].number_format = FMT_KRW


# ─────────────── 시트 2: ★종목손익현황 ───────────────
def _build_stock_pnl_sheet(ws, pnl_df, holdings_df, stock_sheet_map, trades=None):
    """종목손익현황 시트 — 한 줄 = 한 종목"""
    ws.sheet_view.showGridLines = False
    
    # 종목별 수수료/거래세 집계 (trades 사용)
    fee_map = defaultdict(lambda: {'수수료': 0, '거래세': 0})
    if trades:
        for t in trades:
            if t.get('거래구분') in ('매수', '매도'):
                name = t.get('종목명', '')
                fee_map[name]['수수료'] += float(t.get('수수료(원)', 0) or 0)
                fee_map[name]['거래세'] += float(t.get('세금(원)', 0) or 0)
    
    # 열 너비
    widths = {'A': 6, 'B': 28, 'C': 8, 'D': 10, 'E': 16, 'F': 12, 'G': 14, 'H': 16,
              'I': 12, 'J': 16, 'K': 14, 'L': 14, 'M': 12, 'N': 16, 'O': 12, 'P': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    
    # 헤더
    headers = ['번호', '종목명', '통화', '증권사',
               '매수금액', '매수수량', '평균단가',
               '매도금액', '매도수량',
               '매도차손익', '처분이익(+)', '처분손실(-)',
               '보유수량', '평가금액', '수수료', '거래세']
    
    # 제목
    ws['A1'] = '★ 종목별 손익 현황'
    ws['A1'].font = FONT_TITLE
    ws.merge_cells('A1:P1')
    
    ws['A2'] = f'총 {len(pnl_df)}개 종목 · 작성일자: {datetime.now().strftime("%Y-%m-%d")}'
    ws['A2'].font = FONT_SUBTITLE
    ws.merge_cells('A2:P2')
    
    # 헤더 행 (3행)
    for i, h in enumerate(headers):
        c = ws.cell(row=3, column=i+1, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    
    # 데이터 (4행부터)
    if pnl_df.empty:
        return
    
    # 보유 종목 매핑 (보유수량/평가금액)
    holdings_map = {}
    if not holdings_df.empty:
        for _, row in holdings_df.iterrows():
            key = row['종목명']
            holdings_map[key] = {
                '보유수량': row['보유수량'],
                '평가금액': row.get('평가금액(원)', 0) or 0,
            }
    
    for i, (_, row) in enumerate(pnl_df.iterrows()):
        r = i + 4
        stock_name = row['종목명']
        sheet_name = stock_sheet_map.get(stock_name, '')
        
        ws.cell(row=r, column=1, value=i+1)
        
        # 종목명 + 하이퍼링크
        if sheet_name:
            ws.cell(row=r, column=2, value=f"=HYPERLINK(\"#'{sheet_name}'!A1\",\"{stock_name}\")")
            ws.cell(row=r, column=2).font = FONT_LINK
        else:
            ws.cell(row=r, column=2, value=stock_name)
            ws.cell(row=r, column=2).font = FONT_BODY
        
        ws.cell(row=r, column=3, value=row['통화'])
        ws.cell(row=r, column=4, value=row['증권사'])
        
        # 매수
        ws.cell(row=r, column=5, value=float(row['총매수(원)']))
        ws.cell(row=r, column=6, value=float(row['매수수량']))
        ws.cell(row=r, column=7, value=float(row['평균매수단가(원)']))
        
        # 매도
        ws.cell(row=r, column=8, value=float(row['총매도(원)']))
        ws.cell(row=r, column=9, value=float(row['매도수량']))
        
        # 손익
        ws.cell(row=r, column=10, value=float(row['처분손익(원)']))
        ws.cell(row=r, column=11, value=float(row['처분이익(+)']))
        ws.cell(row=r, column=12, value=float(row.get('처분손실(-)', 0)))
        
        # 잔고/평가
        hold = holdings_map.get(stock_name, {})
        if hold:
            ws.cell(row=r, column=13, value=float(hold['보유수량']))
            ws.cell(row=r, column=14, value=float(hold['평가금액']))
        
        # 수수료/거래세 — trades에서 집계한 값 사용
        fees = fee_map.get(stock_name, {'수수료': 0, '거래세': 0})
        ws.cell(row=r, column=15, value=fees['수수료'])
        ws.cell(row=r, column=16, value=fees['거래세'])
        
        # 서식
        for col_idx in range(1, 17):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BORDER_ALL
            if col_idx == 1:
                cell.alignment = ALIGN_CENTER
                cell.font = FONT_BODY
            elif col_idx == 2:
                cell.alignment = ALIGN_LEFT
            elif col_idx in (3, 4):
                cell.alignment = ALIGN_CENTER
                cell.font = FONT_BODY
            else:
                cell.alignment = ALIGN_RIGHT
                cell.font = FONT_BODY
                if col_idx in (6, 9, 13):  # 수량
                    cell.number_format = '#,##0.####'
                else:
                    cell.number_format = FMT_KRW
        
        ws.row_dimensions[r].height = 18
    
    # 합계 행
    total_r = len(pnl_df) + 4
    ws.cell(row=total_r, column=1, value='').fill = FILL_TOTAL
    ws.cell(row=total_r, column=2, value='합계').font = FONT_BODY_BOLD
    ws.cell(row=total_r, column=2).fill = FILL_TOTAL
    ws.cell(row=total_r, column=2).alignment = ALIGN_LEFT
    ws.merge_cells(start_row=total_r, start_column=2, end_row=total_r, end_column=4)
    
    for col_idx in [5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16]:
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=total_r, column=col_idx, value=f'=SUM({col_letter}4:{col_letter}{total_r-1})')
        cell.fill = FILL_TOTAL
        cell.font = FONT_BODY_BOLD
        cell.alignment = ALIGN_RIGHT
        cell.border = BORDER_ALL
        if col_idx in (6, 9, 13):
            cell.number_format = '#,##0.####'
        else:
            cell.number_format = FMT_KRW
    
    for col_idx in [1, 3, 4]:
        ws.cell(row=total_r, column=col_idx).fill = FILL_TOTAL
        ws.cell(row=total_r, column=col_idx).border = BORDER_ALL
    
    # 틀 고정 (헤더)
    ws.freeze_panes = 'C4'


# ─────────────── 시트 3: 종목별 거래내역 ───────────────
def _build_stock_detail_sheet(ws, stock_name, trades, pnl_df, holdings_df):
    """종목별 거래내역 시트 — 사장님 양식 (한 종목당 한 시트)"""
    ws.sheet_view.showGridLines = False
    
    # 열 너비
    widths = {'A': 14, 'B': 10, 'C': 8, 'D': 10, 'E': 14, 'F': 14, 'G': 14,
              'H': 14, 'I': 14, 'J': 12, 'K': 12, 'L': 16, 'M': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    
    # 해당 종목 거래만 필터
    stock_trades = [t for t in trades 
                    if t.get('종목명') == stock_name 
                    and t.get('거래구분') in ('매수', '매도', '배당', '분배금')]
    
    # 제목 + 네비
    ws['A1'] = f'📊 {stock_name}'
    ws['A1'].font = FONT_TITLE
    ws.merge_cells('A1:M1')
    
    ws['A2'] = '=HYPERLINK("#\'★투자성과 대시보드\'!A1","← 대시보드로 돌아가기")'
    ws['A2'].font = FONT_LINK
    
    ws['F2'] = '=HYPERLINK("#\'★종목손익현황\'!A1","← 종목손익현황으로")'
    ws['F2'].font = FONT_LINK
    
    # 헤더 (4행)
    headers = ['거래일자', '증권사', '통화', '거래구분',
               '수량', '단가', '거래금액(외화)', '환율', '원화환산',
               '수수료', '거래세', '비고', '누적잔고']
    
    for i, h in enumerate(headers):
        c = ws.cell(row=4, column=i+1, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    
    # 데이터
    running_qty = 0
    for i, t in enumerate(sorted(stock_trades, key=lambda x: x['거래일자'])):
        r = i + 5
        
        qty = t.get('수량', 0) or 0
        if t['거래구분'] == '매수':
            running_qty += qty
        elif t['거래구분'] == '매도':
            running_qty -= qty
        
        ws.cell(row=r, column=1, value=t.get('거래일자', ''))
        ws.cell(row=r, column=2, value=t.get('증권사', ''))
        ws.cell(row=r, column=3, value=t.get('통화', ''))
        ws.cell(row=r, column=4, value=t.get('거래구분', ''))
        ws.cell(row=r, column=5, value=float(qty))
        ws.cell(row=r, column=6, value=float(t.get('단가', 0) or 0))
        ws.cell(row=r, column=7, value=float(t.get('거래금액', 0) or 0))
        ws.cell(row=r, column=8, value=float(t.get('환율', 0) or 0))
        ws.cell(row=r, column=9, value=float(t.get('원화환산금액', 0) or 0))
        ws.cell(row=r, column=10, value=float(t.get('수수료(원)', 0) or 0))
        ws.cell(row=r, column=11, value=float(t.get('세금(원)', 0) or 0))
        ws.cell(row=r, column=12, value=str(t.get('비고', '') or ''))
        ws.cell(row=r, column=13, value=running_qty if t['거래구분'] in ('매수', '매도') else None)
        
        # 서식
        for col_idx in range(1, 14):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BORDER_ALL
            cell.font = FONT_BODY
            if col_idx in (1, 2, 3, 4):
                cell.alignment = ALIGN_CENTER
            elif col_idx == 12:
                cell.alignment = ALIGN_LEFT
            else:
                cell.alignment = ALIGN_RIGHT
                if col_idx == 5 or col_idx == 13:
                    cell.number_format = '#,##0.####'
                elif col_idx == 6 or col_idx == 8:
                    cell.number_format = '#,##0.##'
                else:
                    cell.number_format = FMT_KRW
            
            # 매도는 옅은 회색 배경
            if t['거래구분'] == '매도':
                cell.fill = PatternFill('solid', fgColor='FAFAFA')
        
        ws.row_dimensions[r].height = 18
    
    # 합계 행
    if stock_trades:
        total_r = len(stock_trades) + 5
        ws.cell(row=total_r, column=1, value='합계').font = FONT_BODY_BOLD
        ws.cell(row=total_r, column=1).fill = FILL_TOTAL
        ws.cell(row=total_r, column=1).alignment = ALIGN_CENTER
        ws.cell(row=total_r, column=1).border = BORDER_ALL
        
        for col_idx in [5, 7, 9, 10, 11]:
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=total_r, column=col_idx, 
                          value=f'=SUM({col_letter}5:{col_letter}{total_r-1})')
            cell.fill = FILL_TOTAL
            cell.font = FONT_BODY_BOLD
            cell.alignment = ALIGN_RIGHT
            cell.border = BORDER_ALL
            if col_idx == 5:
                cell.number_format = '#,##0.####'
            else:
                cell.number_format = FMT_KRW
        
        for col_idx in [2, 3, 4, 6, 8, 12, 13]:
            ws.cell(row=total_r, column=col_idx).fill = FILL_TOTAL
            ws.cell(row=total_r, column=col_idx).border = BORDER_ALL
    
    # 틀 고정
    ws.freeze_panes = 'A5'
