"""기초잔고 처리 모듈

3가지 입력 방식 지원:
- 옵션 A: 사이드바 표 직접 입력 → DataFrame
- 옵션 B: 엑셀 업로드 → DataFrame
- 옵션 C: 이전 연도 거래내역 → 자동 매수원가 계산

모든 방식의 결과는 calculate_stock_pnl()이 받는 opening_balance 형식으로 변환:
[{'종목코드': str, '종목명': str, '통화': str, '수량': float, '평균단가': float, '원화금액': float}, ...]
"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# 기초잔고 양식 컬럼 (사장님 워크북과 호환)
TEMPLATE_COLUMNS = [
    '종목코드',
    '종목명',
    '통화',
    '보유수량',
    '평균매수단가(원화)',
    '비고',
]


def build_opening_balance_template():
    """기초잔고 입력 템플릿 엑셀 생성 (사용자가 다운받아 채워 다시 업로드)
    
    Returns:
        BytesIO: 엑셀 파일 바이너리
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '기초잔고'
    
    # 스타일
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='BFBFBF')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    body_font = Font(name='Arial', size=10)
    body_align_center = Alignment(horizontal='center', vertical='center')
    body_align_left = Alignment(horizontal='left', vertical='center', indent=1)
    body_align_right = Alignment(horizontal='right', vertical='center')
    
    # 제목
    ws['A1'] = '📋 기초잔고 입력'
    ws['A1'].font = Font(name='Arial', size=14, bold=True, color='1F3864')
    ws.merge_cells('A1:F1')
    
    ws['A2'] = '이전 연도 말 기준 보유 중인 종목을 입력하세요. 양도세 계산이 정확해집니다.'
    ws['A2'].font = Font(name='Arial', size=9, color='666666')
    ws.merge_cells('A2:F2')
    
    # 헤더 (행 4)
    for i, col in enumerate(TEMPLATE_COLUMNS):
        c = ws.cell(row=4, column=i+1, value=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border_all
    
    # 예시 데이터 (참고용)
    examples = [
        ('005380', '현대자동차', 'KRW', 41, 220000, '예시 1 — 국내 KOSPI'),
        ('086790', '하나금융지주', 'KRW', 161, 56800, '예시 2 — 국내 KOSPI'),
        ('US23804L1035', '데이터독', 'USD', 5, 215000, '예시 3 — 해외주식 (원화 환산 단가)'),
    ]
    
    for i, row_data in enumerate(examples):
        r = 5 + i
        for j, val in enumerate(row_data):
            cell = ws.cell(row=r, column=j+1, value=val)
            cell.font = body_font
            cell.border = border_all
            if j == 0:
                cell.alignment = body_align_center
            elif j == 1:
                cell.alignment = body_align_left
            elif j == 2:
                cell.alignment = body_align_center
            elif j in (3, 4):
                cell.alignment = body_align_right
                if j == 3:
                    cell.number_format = '#,##0.####'
                else:
                    cell.number_format = '#,##0'
            else:
                cell.alignment = body_align_left
        # 예시 행은 옅은 노란 배경
        for j in range(1, 7):
            ws.cell(row=r, column=j).fill = PatternFill('solid', fgColor='FFF9E6')
    
    # 빈 입력 행 30개
    for i in range(8, 38):
        for j in range(1, 7):
            cell = ws.cell(row=i, column=j)
            cell.border = border_all
            if j == 0:
                cell.alignment = body_align_center
            elif j in (3, 4):
                cell.alignment = body_align_right
                if j == 3:
                    cell.number_format = '#,##0.####'
                else:
                    cell.number_format = '#,##0'
    
    # 열 너비
    widths = {'A': 18, 'B': 30, 'C': 8, 'D': 12, 'E': 18, 'F': 35}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    
    # 안내 시트 추가
    ws2 = wb.create_sheet('📖 작성 가이드')
    guide_lines = [
        ('📋 기초잔고 작성 가이드', 14, True),
        ('', 9, False),
        ('▶ 종목코드', 11, True),
        ('  · 국내: 6자리 숫자 (예: 005380 현대자동차)', 10, False),
        ('  · 해외: ISIN 12자리 (예: US23804L1035 데이터독)', 10, False),
        ('  · 모르면 비워두셔도 됩니다 (종목명으로만 매칭)', 10, False),
        ('', 9, False),
        ('▶ 통화', 11, True),
        ('  · KRW (한국 원) / USD (미국 달러) / JPY (일본 엔) / HKD (홍콩 달러) / CNY (중국 위안)', 10, False),
        ('', 9, False),
        ('▶ 보유수량', 11, True),
        ('  · 이전 연도 12월 31일 기준 보유 수량', 10, False),
        ('  · 정수 또는 소수 (해외주식 소수점 매수도 가능)', 10, False),
        ('', 9, False),
        ('▶ 평균매수단가(원화)', 11, True),
        ('  · 1주당 평균 매수가 (원화 기준)', 10, False),
        ('  · 해외주식도 매수 시 환율로 환산한 원화 단가', 10, False),
        ('  · 증권사 잔고 화면에서 "평균단가" 또는 "매입가" 확인', 10, False),
        ('', 9, False),
        ('💡 작성 팁', 11, True),
        ('  · 위 예시 행은 지우고 사용하세요 (실제 데이터 입력 시)', 10, False),
        ('  · 행이 부족하면 빈 행 아래에 그냥 이어서 입력', 10, False),
        ('  · 종목코드는 비워둬도 OK, 종목명은 필수', 10, False),
        ('  · 평균단가는 1주당 단가입니다 (총 매수금액 ÷ 수량)', 10, False),
    ]
    for i, (text, size, bold) in enumerate(guide_lines):
        cell = ws2.cell(row=i+1, column=1, value=text)
        cell.font = Font(name='Arial', size=size, bold=bold,
                         color='1F3864' if bold and size >= 11 else '000000')
    ws2.column_dimensions['A'].width = 80
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def parse_opening_balance_excel(file_obj):
    """기초잔고 엑셀 → opening_balance 리스트 변환
    
    Args:
        file_obj: 엑셀 파일 (Streamlit uploaded_file 또는 경로)
    
    Returns:
        list of dict: opening_balance 형식
    """
    # 헤더 자동 감지 (첫 5행 안에서 찾기)
    if hasattr(file_obj, 'seek'):
        file_obj.seek(0)
    
    df_raw = pd.read_excel(file_obj, sheet_name=0, header=None, engine='openpyxl')
    
    header_row = None
    for i in range(min(10, len(df_raw))):
        row_vals = [str(v).strip() for v in df_raw.iloc[i].tolist() if pd.notna(v)]
        # '종목명' + '수량' 또는 '보유수량' 둘 다 있으면 헤더
        joined = ' '.join(row_vals)
        if '종목명' in joined and ('수량' in joined):
            header_row = i
            break
    
    if header_row is None:
        raise ValueError("기초잔고 양식이 아닙니다. '종목명'과 '보유수량' 컬럼이 필요합니다.")
    
    # 헤더 행으로 다시 읽기
    if hasattr(file_obj, 'seek'):
        file_obj.seek(0)
    df = pd.read_excel(file_obj, sheet_name=0, header=header_row, engine='openpyxl')
    
    # 컬럼 매핑 (유연하게)
    col_map = {}
    for col in df.columns:
        col_s = str(col).strip()
        if col_s in ('종목코드', '코드', 'code', 'ISIN'):
            col_map['code'] = col
        elif col_s in ('종목명', '종목', 'name'):
            col_map['name'] = col
        elif col_s in ('통화', '통화코드', 'currency'):
            col_map['currency'] = col
        elif col_s in ('보유수량', '수량', '잔고수량', 'qty'):
            col_map['qty'] = col
        elif col_s in ('평균매수단가(원화)', '평균매수단가', '평균단가', '평단', '평단가', 'avg_price'):
            col_map['price'] = col
    
    if 'name' not in col_map or 'qty' not in col_map:
        raise ValueError(f"필수 컬럼 누락: 종목명, 보유수량. 발견된 컬럼: {list(df.columns)}")
    
    # 예시 행은 노란 배경으로 표시했지만 색깔로는 못 거르니 비고에 '예시'있는 행 거르기
    opening = []
    for _, row in df.iterrows():
        name = row.get(col_map['name'])
        if pd.isna(name) or not str(name).strip():
            continue
        name = str(name).strip()
        
        # 비고에 '예시' 들어있으면 스킵
        비고 = ''
        if '비고' in df.columns:
            v = row.get('비고')
            if pd.notna(v):
                비고 = str(v)
                if '예시' in 비고:
                    continue
        
        qty = row.get(col_map['qty'])
        if pd.isna(qty):
            continue
        try:
            qty = float(qty)
        except (ValueError, TypeError):
            continue
        if qty <= 0:
            continue
        
        # 평균단가
        price = 0
        if 'price' in col_map:
            p = row.get(col_map['price'])
            if pd.notna(p):
                try:
                    price = float(p)
                except (ValueError, TypeError):
                    price = 0
        
        code = ''
        if 'code' in col_map:
            c = row.get(col_map['code'])
            if pd.notna(c):
                code = str(c).strip()
                # 숫자형 종목코드가 float으로 읽힐 수 있음 (예: 5380.0 → "005380")
                if code.endswith('.0'):
                    code = code[:-2]
                if code.isdigit() and len(code) < 6:
                    code = code.zfill(6)
        
        currency = 'KRW'
        if 'currency' in col_map:
            c = row.get(col_map['currency'])
            if pd.notna(c):
                currency = str(c).strip().upper()
        
        opening.append({
            '종목코드': code,
            '종목명': name,
            '통화': currency,
            '수량': qty,
            '평균단가': price,
            '원화금액': qty * price,
        })
    
    return opening


def dataframe_to_opening_balance(df):
    """사이드바 표(data_editor) DataFrame → opening_balance 리스트 변환
    
    Args:
        df: pandas DataFrame (TEMPLATE_COLUMNS 양식)
    
    Returns:
        list of dict
    """
    if df is None or df.empty:
        return []
    
    opening = []
    for _, row in df.iterrows():
        name = row.get('종목명')
        if pd.isna(name) or not str(name).strip():
            continue
        name = str(name).strip()
        
        qty = row.get('보유수량')
        if pd.isna(qty):
            continue
        try:
            qty = float(qty)
        except (ValueError, TypeError):
            continue
        if qty <= 0:
            continue
        
        price = 0
        p = row.get('평균매수단가(원화)')
        if pd.notna(p):
            try:
                price = float(p)
            except (ValueError, TypeError):
                price = 0
        
        code = ''
        c = row.get('종목코드')
        if pd.notna(c):
            code = str(c).strip()
            if code.endswith('.0'):
                code = code[:-2]
            if code.isdigit() and len(code) < 6:
                code = code.zfill(6)
        
        currency = 'KRW'
        cc = row.get('통화')
        if pd.notna(cc):
            currency = str(cc).strip().upper()
        
        opening.append({
            '종목코드': code,
            '종목명': name,
            '통화': currency,
            '수량': qty,
            '평균단가': price,
            '원화금액': qty * price,
        })
    
    return opening


def get_empty_template_df():
    """data_editor용 빈 DataFrame (5행 미리 준비)"""
    return pd.DataFrame({
        '종목코드': ['', '', '', '', ''],
        '종목명': ['', '', '', '', ''],
        '통화': ['KRW', 'KRW', 'KRW', 'KRW', 'KRW'],
        '보유수량': [None, None, None, None, None],
        '평균매수단가(원화)': [None, None, None, None, None],
    })
